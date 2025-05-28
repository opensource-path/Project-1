import os
from openpyxl import load_workbook, Workbook

# Directory containing the Excel files
input_dir = r"C:\Users\Julien.Whitter\Downloads\Temporary Files\State Sheets\IN\Downloads"
output_file = r"C:\Users\Julien.Whitter\Downloads\Temporary Files\State Sheets\IN\Combined_Sheets.xlsx"

# Create a new workbook for the combined data
combined_workbook = Workbook()

# Remove the default sheet created in the new workbook
if "Sheet" in combined_workbook.sheetnames:
    del combined_workbook["Sheet"]

# Iterate over Excel files in the directory
for file_name in os.listdir(input_dir):
    if file_name.endswith(".xlsx") or file_name.endswith(".xlsm"):  # Include Excel files only
        file_path = os.path.join(input_dir, file_name)
        try:
            # Load the workbook and select the 7th sheet
            workbook = load_workbook(file_path, data_only=True)
            sheet_names = workbook.sheetnames
            if len(sheet_names) >= 7:
                sheet_name = sheet_names[6]  # Get the 7th sheet (index 6)
                sheet = workbook[sheet_name]

                # Create a new sheet in the combined workbook
                combined_sheet_name = file_name.split(".")[0][:31]  # Limit to 31 characters for Excel sheet names
                combined_worksheet = combined_workbook.create_sheet(title=combined_sheet_name)

                # Copy data from the 7th sheet to the new sheet in the combined workbook
                for row in sheet.iter_rows(values_only=True):
                    combined_worksheet.append(row)
            else:
                print(f"{file_name} does not have 7 sheets. Skipping...")
        except Exception as e:
            print(f"Error processing file {file_name}: {e}")

# Save the combined workbook
combined_workbook.save(output_file)
print(f"Combined workbook saved as {output_file}")
