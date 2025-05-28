import os
from openpyxl import load_workbook, Workbook

# Directory containing the Excel files
input_dir = r"C:\Users\Julien.Whitter\Downloads\Temporary Files\State Sheets\IN\Downloads"
output_file = r"C:\Users\Julien.Whitter\Downloads\Temporary Files\State Sheets\IN\Combined_Sheets_With_Tabs.xlsx"

# Create a new workbook for the combined data
combined_workbook = Workbook()
# Remove the default sheet created in the new workbook
if "Sheet" in combined_workbook.sheetnames:
    default_sheet = combined_workbook["Sheet"]
    combined_workbook.remove(default_sheet)

# Iterate over Excel files in the directory
for file_name in os.listdir(input_dir):
    if file_name.endswith(".xlsx") or file_name.endswith(".xlsm"):  # Include Excel files only
        file_path = os.path.join(input_dir, file_name)
        try:
            # Load the workbook and select the 8th sheet
            workbook = load_workbook(file_path, data_only=True)
            sheet_names = workbook.sheetnames
            if len(sheet_names) >= 8:
                sheet_name = sheet_names[7]  # Get the 8th sheet (index 7)
                if sheet_name.strip() in ["Sheet8", "Sheet8 "]:
                    sheet = workbook[sheet_name]

                    # Create a new sheet in the combined workbook with the workbook name as the tab name
                    new_sheet_name = os.path.splitext(file_name)[0]  # Remove file extension
                    new_sheet = combined_workbook.create_sheet(title=new_sheet_name)

                    # Copy data from the 8th sheet to the new sheet
                    for row in sheet.iter_rows(values_only=True):
                        new_sheet.append(row)
            else:
                print(f"{file_name} does not have 8 sheets. Skipping...")
        except Exception as e:
            print(f"Error processing file {file_name}: {e}")

# Save the combined workbook
combined_workbook.save(output_file)
print(f"Combined workbook saved as {output_file}")
